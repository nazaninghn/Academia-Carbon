from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseForbidden
from .models import Country, EmissionData, EmissionRecord, Supplier, MaterialRequest
from django.db.models import Sum, Avg, Count, FloatField
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.contrib import messages
from .forms import EmailLoginForm, EmailSignupForm
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_http_methods
import logging
import json

# Import Arcjet simulation for enhanced security
from .arcjet_simulation import arcjet_protect

# PHASE 3 ‚Äî RATE LIMIT (8Ô∏è‚É£ ÿ¨ŸÑŸà⁄Ø€åÿ±€å ÿßÿ≤ brute-force)
try:
    from django_ratelimit.decorators import ratelimit
    RATELIMIT_AVAILABLE = True
except ImportError:
    # Fallback decorator if ratelimit not installed
    def ratelimit(key=None, rate=None, method=None, block=False):
        def decorator(func):
            return func
        return decorator
    RATELIMIT_AVAILABLE = False

# Security logger
security_logger = logging.getLogger('ghg.security')

def index(request):
    # Redirect to login if not authenticated
    if not request.user.is_authenticated:
        return redirect('ghg:email_login')
    
    countries = Country.objects.all().order_by('name')
    latest_year = EmissionData.objects.order_by('-year').first()
    total_emissions = EmissionData.objects.filter(
        year=latest_year.year if latest_year else 2023
    ).aggregate(Sum('total_ghg'))['total_ghg__sum'] or 0
    
    data_points = EmissionData.objects.count()
    
    context = {
        'countries': countries,
        'total_emissions': round(total_emissions, 2),
        'latest_year': latest_year.year if latest_year else 2023,
        'data_points': data_points,
        'active_menu': 'dashboard',  # PHASE 5 ‚Äî UI/UX (13Ô∏è‚É£ ŸÖÿ¥⁄©ŸÑ ÿ™ÿ∫€å€åÿ± ÿØÿßÿ¥ÿ®Ÿàÿ±ÿØ)
    }
    return render(request, 'index.html', context)

# PHASE 2 ‚Äî AUTH & PERMISSIONS (5Ô∏è‚É£ ŸáŸÖŸá viewŸáÿß€å ÿ≠ÿ≥ÿßÿ≥ login_required)
@login_required
def data_entry(request):
    context = {
        'active_menu': 'emission_management',  # PHASE 5 ‚Äî UI/UX (13Ô∏è‚É£)
    }
    return render(request, 'data_entry.html', context)

# PHASE 2 ‚Äî AUTH & PERMISSIONS + PHASE 3 ‚Äî RATE LIMIT
@login_required
@csrf_protect
@require_http_methods(["POST"])
@ratelimit(key='user', rate='30/m', method='POST', block=True)
@arcjet_protect()  # Enhanced security with Arcjet simulation
def calculate_emission(request):
    """Calculate emissions with security checks"""
    try:
        from .emission_factors import calculate_emissions
        from .models import EmissionRecord, Supplier
        import logging
        logger = logging.getLogger(__name__)
        
        data = json.loads(request.body)
        category = data.get('category')
        source = data.get('source')
        activity_data = float(data.get('activity_data', 0))
        country = data.get('country', 'global')
        description = data.get('description', '')
        industry_type = data.get('industry_type', '')
        fuel_name = data.get('fuel_name', '')
        supplier_id = data.get('supplier_id', None)
        save_record = data.get('save', True)
        
        # Debug logging
        logger.info(f"Received calculation request - industry_type: '{industry_type}', supplier_id: '{supplier_id}'")
        
        # Security: Validate input data
        if activity_data < 0 or activity_data > 1000000:
            security_logger.warning(f"Suspicious activity data: {activity_data} from user {request.user.id}")
            return JsonResponse({'error': 'Invalid activity data'}, status=400)
        
        result = calculate_emissions(category, source, activity_data, country)
        
        if 'error' not in result and save_record:
            # Determine scope based on category
            scope = '1'  # Default
            if category in ['electricity', 'steam-heat']:
                scope = '2'
            elif category in ['travel', 'waste', 'water', 'purchased-goods', 'capital-goods', 
                             'fuel-energy', 'upstream-transport', 'commuting', 'upstream-leased',
                             'downstream-transport', 'end-of-life', 'franchises', 'investments']:
                scope = '3'
            
            # PHASE 2 ‚Äî AUTH & PERMISSIONS (6Ô∏è‚É£ ŸÅ€åŸÑÿ™ÿ± ÿØÿßÿØŸá ÿ®ÿ± ÿßÿ≥ÿßÿ≥ user)
            supplier_obj = None
            if supplier_id:
                try:
                    # ‚úÖ ÿØÿ±ÿ≥ÿ™: ŸÅ€åŸÑÿ™ÿ± ÿ®ÿ± ÿßÿ≥ÿßÿ≥ user
                    supplier_obj = Supplier.objects.get(id=supplier_id, user=request.user)
                except Supplier.DoesNotExist:
                    security_logger.warning(f"User {request.user.id} tried to access supplier {supplier_id}")
                    pass
            
            # Save to database with user association
            record = EmissionRecord.objects.create(
                user=request.user,  # Always associate with current user
                scope=scope,
                category=category,
                source=source,
                source_name=result['source_name'],
                activity_data=activity_data,
                unit=result['unit'],
                emission_factor=result['factor'],
                emissions_kg=result['emissions_kg'],
                emissions_tons=result['emissions_tons'],
                country=country,
                reference=result.get('reference', ''),
                description=description[:500],  # Limit description length
                industry_type=industry_type[:100] if industry_type else None,
                fuel_name=fuel_name[:100] if fuel_name else None,
                supplier=supplier_obj
            )
            
            result['record_id'] = record.id
            result['saved'] = True
            
            # Log successful emission calculation
            security_logger.info(f"Emission calculated by user {request.user.id}: {result['emissions_kg']} kg CO2e")
        
        return JsonResponse(result)
        
    except json.JSONDecodeError:
        security_logger.warning(f"Invalid JSON from user {request.user.id}")
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except ValueError as e:
        security_logger.warning(f"Invalid data from user {request.user.id}: {str(e)}")
        return JsonResponse({'error': 'Invalid input data'}, status=400)
    except Exception as e:
        security_logger.error(f"Emission calculation error for user {request.user.id}: {str(e)}")
        return JsonResponse({'error': 'Calculation failed'}, status=500)

# PHASE 2 ‚Äî AUTH & PERMISSIONS (5Ô∏è‚É£, 6Ô∏è‚É£, 7Ô∏è‚É£)
@login_required
@ratelimit(key='user', rate='10/m', method='GET', block=True)
def get_emission_records(request):
    """Get user's emission records with pagination"""
    try:
        # PHASE 2 ‚Äî AUTH & PERMISSIONS (6Ô∏è‚É£ ŸÅ€åŸÑÿ™ÿ± ÿØÿßÿØŸá ÿ®ÿ± ÿßÿ≥ÿßÿ≥ user)
        # ‚úÖ ÿØÿ±ÿ≥ÿ™: ŸÅŸÇÿ∑ ÿ±⁄©Ÿàÿ±ÿØŸáÿß€å ⁄©ÿßÿ±ÿ®ÿ± ŸÅÿπŸÑ€å
        records = EmissionRecord.objects.filter(user=request.user).order_by('-created_at')
        
        # Pagination
        page = int(request.GET.get('page', 1))
        per_page = min(int(request.GET.get('per_page', 20)), 100)  # Max 100 per page
        
        start = (page - 1) * per_page
        end = start + per_page
        
        records_page = records[start:end]
        
        data = []
        for record in records_page:
            data.append({
                'id': record.id,
                'scope': record.scope,
                'category': record.category,
                'source_name': record.source_name,
                'activity_data': record.activity_data,
                'unit': record.unit,
                'emissions_kg': record.emissions_kg,
                'emissions_tons': record.emissions_tons,
                'created_at': record.created_at.isoformat(),
                'description': record.description,
            })
        
        return JsonResponse({
            'records': data,
            'total': records.count(),
            'page': page,
            'per_page': per_page,
        })
        
    except Exception as e:
        security_logger.error(f"Error fetching records for user {request.user.id}: {str(e)}")
        return JsonResponse({'error': 'Failed to fetch records'}, status=500)

# PHASE 2 ‚Äî AUTH & PERMISSIONS (7Ô∏è‚É£ ÿ¨ŸÑŸà⁄Ø€åÿ±€å ÿßÿ≤ ÿØÿ≥ÿ™ÿ±ÿ≥€å URL ÿØÿ≥ÿ™€å)
@login_required
@csrf_protect
@require_http_methods(["DELETE"])
@ratelimit(key='user', rate='20/m', method='DELETE', block=True)
def delete_emission_record(request, record_id):
    """Delete emission record with security checks"""
    try:
        # PHASE 2 ‚Äî AUTH & PERMISSIONS (6Ô∏è‚É£, 7Ô∏è‚É£)
        # ‚úÖ ÿØÿ±ÿ≥ÿ™: ŸÅŸÇÿ∑ ÿ±⁄©Ÿàÿ±ÿØŸáÿß€å ⁄©ÿßÿ±ÿ®ÿ± ŸÅÿπŸÑ€å + ÿ®ÿ±ÿ±ÿ≥€å ŸÖÿßŸÑ⁄©€åÿ™
        record = get_object_or_404(EmissionRecord, id=record_id, user=request.user)
        
        # PHASE 2 ‚Äî AUTH & PERMISSIONS (7Ô∏è‚É£ ÿ¨ŸÑŸà⁄Ø€åÿ±€å ÿßÿ≤ ÿØÿ≥ÿ™ÿ±ÿ≥€å URL ÿØÿ≥ÿ™€å)
        if record.user != request.user:
            security_logger.warning(f"User {request.user.id} tried to delete record {record_id} owned by {record.user.id}")
            return HttpResponseForbidden("Access denied")
        
        record.delete()
        
        security_logger.info(f"User {request.user.id} deleted emission record {record_id}")
        
        return JsonResponse({'success': True})
        
    except EmissionRecord.DoesNotExist:
        security_logger.warning(f"User {request.user.id} tried to delete non-existent record {record_id}")
        return JsonResponse({'error': 'Record not found'}, status=404)
    except Exception as e:
        security_logger.error(f"Error deleting record {record_id} for user {request.user.id}: {str(e)}")
        return JsonResponse({'error': 'Failed to delete record'}, status=500)

# PHASE 2 ‚Äî AUTH & PERMISSIONS (6Ô∏è‚É£ ŸÅ€åŸÑÿ™ÿ± ÿØÿßÿØŸá ÿ®ÿ± ÿßÿ≥ÿßÿ≥ user)
@login_required
@ratelimit(key='user', rate='10/m', method='GET', block=True)
def get_suppliers(request):
    """Get user's suppliers"""
    try:
        # ‚úÖ ÿØÿ±ÿ≥ÿ™: ŸÅŸÇÿ∑ ÿ™ÿßŸÖ€åŸÜ‚Äå⁄©ŸÜŸÜÿØ⁄ØÿßŸÜ ⁄©ÿßÿ±ÿ®ÿ± ŸÅÿπŸÑ€å
        suppliers = Supplier.objects.filter(user=request.user).order_by('name')
        
        data = []
        for supplier in suppliers:
            data.append({
                'id': supplier.id,
                'name': supplier.name,
                'supplier_type': supplier.supplier_type,
                'country': supplier.country,
                'email': supplier.email,
            })
        
        return JsonResponse({'suppliers': data})
        
    except Exception as e:
        security_logger.error(f"Error fetching suppliers for user {request.user.id}: {str(e)}")
        return JsonResponse({'error': 'Failed to fetch suppliers'}, status=500)

# PHASE 2 ‚Äî AUTH & PERMISSIONS + PHASE 3 ‚Äî RATE LIMIT
@login_required
@csrf_protect
@require_http_methods(["POST"])
@ratelimit(key='user', rate='10/m', method='POST', block=True)
@arcjet_protect()  # Enhanced security with Arcjet simulation
def add_supplier(request):
    """Add new supplier with validation"""
    try:
        data = json.loads(request.body)
        
        # Security: Validate and sanitize input
        name = data.get('name', '').strip()[:200]
        supplier_type = data.get('supplier_type', '').strip()[:100]
        country = data.get('country', '').strip()[:100]
        email = data.get('email', '').strip()[:254]
        
        if not name:
            return JsonResponse({'error': 'Supplier name is required'}, status=400)
        
        # PHASE 2 ‚Äî AUTH & PERMISSIONS (6Ô∏è‚É£ ŸÅ€åŸÑÿ™ÿ± ÿØÿßÿØŸá ÿ®ÿ± ÿßÿ≥ÿßÿ≥ user)
        # Check for duplicate supplier name for this user
        if Supplier.objects.filter(user=request.user, name=name).exists():
            return JsonResponse({'error': 'Supplier with this name already exists'}, status=400)
        
        supplier = Supplier.objects.create(
            user=request.user,  # Always associate with current user
            name=name,
            supplier_type=supplier_type,
            country=country,
            email=email,
        )
        
        security_logger.info(f"User {request.user.id} added supplier: {name}")
        
        return JsonResponse({
            'success': True,
            'supplier': {
                'id': supplier.id,
                'name': supplier.name,
                'supplier_type': supplier.supplier_type,
                'country': supplier.country,
                'email': supplier.email,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        security_logger.error(f"Error adding supplier for user {request.user.id}: {str(e)}")
        return JsonResponse({'error': 'Failed to add supplier'}, status=500)

# PHASE 3 ‚Äî RATE LIMIT (8Ô∏è‚É£ ÿ¨ŸÑŸà⁄Ø€åÿ±€å ÿßÿ≤ brute-force)
@ratelimit(key="ip", rate="5/m", block=True)
@arcjet_protect()  # Enhanced security with Arcjet simulation
def login_view(request):
    """Login with rate limiting"""
    if request.method == 'POST':
        form = EmailLoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            
            user = authenticate(request, username=email, password=password)
            if user is not None:
                login(request, user)
                security_logger.info(f"Successful login for user: {email}")
                return redirect('ghg:index')
            else:
                security_logger.warning(f"Failed login attempt for email: {email} from IP: {request.META.get('REMOTE_ADDR')}")
                messages.error(request, 'Invalid email or password.')
    else:
        form = EmailLoginForm()
    
    return render(request, 'registration/login.html', {'form': form})

# PHASE 2 ‚Äî AUTH & PERMISSIONS (5Ô∏è‚É£ ŸáŸÖŸá viewŸáÿß€å ÿ≠ÿ≥ÿßÿ≥ login_required)
@login_required
def emissions_view(request):
    """Emissions page with user data isolation"""
    context = {
        'active_menu': 'emissions',  # PHASE 5 ‚Äî UI/UX (13Ô∏è‚É£)
    }
    return render(request, 'emissions.html', context)

@login_required
def settings_view(request):
    """Settings page"""
    context = {
        'active_menu': 'settings',  # PHASE 5 ‚Äî UI/UX (13Ô∏è‚É£)
    }
    return render(request, 'settings.html', context)

@login_required
def suppliers_view(request):
    """Suppliers management page"""
    context = {
        'active_menu': 'suppliers',  # PHASE 5 ‚Äî UI/UX (13Ô∏è‚É£)
    }
    return render(request, 'suppliers.html', context)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def get_country_data(request, country_code):
    try:
        country = Country.objects.get(code=country_code)
        emissions = EmissionData.objects.filter(country=country).order_by('year')
        
        data = {
            'country': country.name,
            'years': [e.year for e in emissions],
            'co2': [e.co2_emissions for e in emissions],
            'methane': [e.methane_emissions or 0 for e in emissions],
            'nitrous_oxide': [e.nitrous_oxide or 0 for e in emissions],
            'total': [e.total_ghg for e in emissions],
        }
        return JsonResponse(data)
    except Country.DoesNotExist:
        return JsonResponse({'error': 'Country not found'}, status=404)

@login_required
def get_global_data(request):
    years = EmissionData.objects.values_list('year', flat=True).distinct().order_by('year')
    
    data = {
        'years': list(years),
        'emissions': []
    }
    
    for year in years:
        total = EmissionData.objects.filter(year=year).aggregate(Sum('total_ghg'))['total_ghg__sum'] or 0
        data['emissions'].append(round(total, 2))
    
    return JsonResponse(data)

@login_required
def get_top_emitters(request):
    # Get latest year
    latest_year = EmissionData.objects.order_by('-year').first()
    year = latest_year.year if latest_year else 2023
    
    # Get top 10 emitters for that year
    top_emitters = EmissionData.objects.filter(year=year).select_related('country').order_by('-total_ghg')[:10]
    
    data = {
        'countries': [e.country.name for e in top_emitters],
        'emissions': [round(e.total_ghg, 2) for e in top_emitters]
    }
    
    return JsonResponse(data)

@login_required
def emissions_summary_api(request):
    """API endpoint for emissions analysis summary data"""
    from django.db.models import Sum, Count
    
    # Get user's emission records
    user_records = EmissionRecord.objects.filter(user=request.user)
    
    # Calculate total emissions in tCO2e
    total_kg = user_records.aggregate(total=Sum('emissions_kg'))['total'] or 0
    total_tco2e = total_kg / 1000  # Convert kg to tonnes
    
    # Count total records
    records_count = user_records.count()
    
    # Standard (always ISO for this system)
    standard = "ISO"
    
    return JsonResponse({
        'total_tco2e': round(total_tco2e, 3),
        'records': records_count,
        'standard': standard
    })


@login_required
def dashboard_api(request):
    """API endpoint for dashboard data"""
    from datetime import datetime, date, timedelta
    from .dashboard_services import get_dashboard_metrics
    from django.db.models import Sum, Count
    from django.db import models
    from decimal import Decimal
    
    # Parse optional filters
    date_from = None
    date_to = None
    country = request.GET.get('country')
    
    if request.GET.get('from'):
        try:
            date_from = datetime.strptime(request.GET.get('from'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if request.GET.get('to'):
        try:
            date_to = datetime.strptime(request.GET.get('to'), '%Y-%m-%d').date()
        except ValueError:
            pass
    
    try:
        # Get comprehensive dashboard metrics
        metrics = get_dashboard_metrics(request.user, date_from, date_to, country)
        
        # Calculate additional metrics for dashboard
        user_records = EmissionRecord.objects.filter(user=request.user)
        
        # Current month emissions
        current_month_start = date.today().replace(day=1)
        current_month_records = user_records.filter(created_at__date__gte=current_month_start)
        current_month_kg = current_month_records.aggregate(total=Sum('emissions_kg', output_field=models.FloatField()))['total'] or 0
        current_month_tons = float(current_month_kg) / 1000
        
        # Previous month for comparison
        if current_month_start.month == 1:
            prev_month_start = current_month_start.replace(year=current_month_start.year - 1, month=12)
            prev_month_end = current_month_start - timedelta(days=1)
        else:
            prev_month_start = current_month_start.replace(month=current_month_start.month - 1)
            prev_month_end = current_month_start - timedelta(days=1)
        
        prev_month_records = user_records.filter(
            created_at__date__gte=prev_month_start,
            created_at__date__lte=prev_month_end
        )
        prev_month_kg = prev_month_records.aggregate(total=Sum('emissions_kg', output_field=models.FloatField()))['total'] or 0
        prev_month_tons = float(prev_month_kg) / 1000
        
        # Calculate percentage changes
        emissions_change_pct = 0
        records_change_pct = 0
        
        if prev_month_tons > 0:
            emissions_change_pct = ((current_month_tons - prev_month_tons) / prev_month_tons) * 100
        
        prev_month_count = prev_month_records.count()
        current_month_count = current_month_records.count()
        if prev_month_count > 0:
            records_change_pct = ((current_month_count - prev_month_count) / prev_month_count) * 100
        
        # Monthly trends for chart (last 6 months)
        monthly_trends = []
        for i in range(6):
            month_date = current_month_start - timedelta(days=30 * i)
            month_start = month_date.replace(day=1)
            
            if month_date.month == 12:
                month_end = month_date.replace(year=month_date.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                month_end = month_date.replace(month=month_date.month + 1, day=1) - timedelta(days=1)
            
            month_records = user_records.filter(
                created_at__date__gte=month_start,
                created_at__date__lte=month_end
            )
            month_kg = month_records.aggregate(total=Sum('emissions_kg', output_field=models.FloatField()))['total'] or 0
            month_tons = float(month_kg) / 1000
            
            monthly_trends.append({
                'month': month_date.strftime('%b %Y'),
                'emissions_tons': round(month_tons, 2)
            })
        
        # Reverse to show chronological order
        monthly_trends.reverse()
        
        # Suppliers count
        suppliers_count = Supplier.objects.filter(user=request.user).count()
        
        # Prepare response data
        response_data = {
            'total_emissions_tons': float(metrics.get('total_emissions_tons', 0)),
            'total_records': metrics.get('total_records', 0),
            'current_month_tons': round(current_month_tons, 2),
            'suppliers_count': suppliers_count,
            'emissions_change_pct': round(emissions_change_pct, 1),
            'records_change_pct': round(records_change_pct, 1),
            'month_change_pct': 0,  # Can be used for other metrics
            'scope_breakdown': metrics.get('scope_breakdown', []),
            'monthly_trends': monthly_trends,
            'latest_records': metrics.get('latest_records', [])[:5],  # Last 5 records
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'error': f'Failed to load dashboard data: {str(e)}',
            'total_emissions_tons': 0,
            'total_records': 0,
            'current_month_tons': 0,
            'suppliers_count': 0,
            'emissions_change_pct': 0,
            'records_change_pct': 0,
            'month_change_pct': 0,
            'scope_breakdown': [],
            'monthly_trends': [],
            'latest_records': []
        }, status=500)
    
    # Get dashboard data
    data = get_dashboard_metrics(
        user=request.user,
        date_from=date_from,
        date_to=date_to,
        country=country,
    )
    
    return JsonResponse(data)

@login_required
def emission_history(request):
    """View to display user's emission calculation history"""
    from .models import EmissionRecord
    
    records = EmissionRecord.objects.filter(user=request.user).order_by('-created_at')[:50]
    
    # Calculate totals by scope
    scope1_total = EmissionRecord.objects.filter(user=request.user, scope='1').aggregate(
        total=Sum('emissions_kg'))['total'] or 0
    scope2_total = EmissionRecord.objects.filter(user=request.user, scope='2').aggregate(
        total=Sum('emissions_kg'))['total'] or 0
    scope3_total = EmissionRecord.objects.filter(user=request.user, scope='3').aggregate(
        total=Sum('emissions_kg'))['total'] or 0
    
    total_emissions = scope1_total + scope2_total + scope3_total
    
    context = {
        'records': records,
        'scope1_total': round(scope1_total / 1000, 2),  # Convert to tons
        'scope2_total': round(scope2_total / 1000, 2),
        'scope3_total': round(scope3_total / 1000, 2),
        'total_emissions': round(total_emissions / 1000, 2),
        'record_count': records.count(),
        'active_menu': 'emission_history',  # Set active menu
    }
    
    return render(request, 'emission_history.html', context)

@login_required
def get_user_emissions_summary(request):
    """API endpoint for user's emission summary"""
    from .models import EmissionRecord
    from django.db.models import Sum
    from datetime import datetime, timedelta
    
    # Get date range (last 12 months)
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365)
    
    records = EmissionRecord.objects.filter(
        user=request.user,
        created_at__gte=start_date
    )
    
    # Group by scope
    scope_data = records.values('scope').annotate(
        total=Sum('emissions_kg')
    ).order_by('scope')
    
    # Group by month
    monthly_data = records.extra(
        select={'month': "strftime('%%Y-%%m', created_at)"}
    ).values('month').annotate(
        total=Sum('emissions_kg')
    ).order_by('month')
    
    # Top categories
    category_data = records.values('category', 'source_name').annotate(
        total=Sum('emissions_kg')
    ).order_by('-total')[:10]
    
    data = {
        'scope_totals': {
            'scope1': round(sum(s['total'] for s in scope_data if s['scope'] == '1') / 1000, 2),
            'scope2': round(sum(s['total'] for s in scope_data if s['scope'] == '2') / 1000, 2),
            'scope3': round(sum(s['total'] for s in scope_data if s['scope'] == '3') / 1000, 2),
        },
        'monthly': [
            {
                'month': m['month'],
                'emissions': round(m['total'] / 1000, 2)
            } for m in monthly_data
        ],
        'top_categories': [
            {
                'name': c['source_name'],
                'emissions': round(c['total'] / 1000, 2)
            } for c in category_data
        ],
        'total_records': records.count()
    }
    
    return JsonResponse(data)

@login_required
def user_guide(request):
    """Display user guide page"""
    context = {
        'active_menu': 'user_guide',  # Set active menu
    }
    return render(request, 'user_guide.html', context)


@login_required
def test_report_feature(request):
    """Test page for report feature"""
    from django.http import HttpResponse
    with open('test_report_feature.html', 'r') as f:
        content = f.read()
    return HttpResponse(content)


@login_required
def test_custom_factor_feature(request):
    """Test page for custom factor feature"""
    from django.shortcuts import render
    return render(request, 'test_custom_factor_simple.html')


@login_required
def report_extra_info_api(request):
    """API endpoint for saving/retrieving report extra information"""
    from .models import ReportExtraInfo
    import json
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Get or create the report extra info for this user
            report_info, created = ReportExtraInfo.objects.get_or_create(
                user=request.user,
                defaults={
                    'legal_name': data.get('legal_name', ''),
                    'industry': data.get('industry', ''),
                    'reporting_period': data.get('period', ''),
                    'boundary_approach': data.get('boundary', ''),
                    'notes': data.get('notes', ''),
                    'share_org_profile': data.get('share_org_profile', True),
                    'share_boundary': data.get('share_boundary', True),
                    'share_data_sources': data.get('share_data_sources', False),
                    'share_projects': data.get('share_projects', False),
                }
            )
            
            # If not created, update existing record
            if not created:
                report_info.legal_name = data.get('legal_name', report_info.legal_name)
                report_info.industry = data.get('industry', report_info.industry)
                report_info.reporting_period = data.get('period', report_info.reporting_period)
                report_info.boundary_approach = data.get('boundary', report_info.boundary_approach)
                report_info.notes = data.get('notes', report_info.notes)
                report_info.share_org_profile = data.get('share_org_profile', report_info.share_org_profile)
                report_info.share_boundary = data.get('share_boundary', report_info.share_boundary)
                report_info.share_data_sources = data.get('share_data_sources', report_info.share_data_sources)
                report_info.share_projects = data.get('share_projects', report_info.share_projects)
                report_info.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Report details saved successfully',
                'created': created
            })
            
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    elif request.method == 'GET':
        try:
            report_info = ReportExtraInfo.objects.get(user=request.user)
            return JsonResponse({
                'legal_name': report_info.legal_name or '',
                'industry': report_info.industry or '',
                'period': report_info.reporting_period or '',
                'boundary': report_info.boundary_approach or '',
                'notes': report_info.notes or '',
                'share_org_profile': report_info.share_org_profile,
                'share_boundary': report_info.share_boundary,
                'share_data_sources': report_info.share_data_sources,
                'share_projects': report_info.share_projects,
            })
        except ReportExtraInfo.DoesNotExist:
            return JsonResponse({
                'legal_name': '',
                'industry': '',
                'period': '',
                'boundary': '',
                'notes': '',
                'share_org_profile': True,
                'share_boundary': True,
                'share_data_sources': False,
                'share_projects': False,
            })
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@arcjet_protect()  # Enhanced security with Arcjet simulation
def email_login_view(request):
    """Login view using email with account lockout protection and reCAPTCHA"""
    from .security import AccountLockout, get_client_ip, log_security_event
    from .captcha import ReCaptchaValidator, get_recaptcha_context
    
    if request.user.is_authenticated:
        return redirect('ghg:index')
    
    # Add reCAPTCHA context
    context = get_recaptcha_context()
    
    if request.method == 'POST':
        form = EmailLoginForm(request.POST)
        email = request.POST.get('username', '')  # username field contains email
        client_ip = get_client_ip(request)
        
        # Validate reCAPTCHA first (if enabled)
        recaptcha_token = request.POST.get('recaptcha_token', '')
        if context.get('RECAPTCHA_ENABLED'):
            if not ReCaptchaValidator.is_human(recaptcha_token, action='login', remote_ip=client_ip):
                messages.error(
                    request,
                    'ü§ñ Security verification failed. Please try again.'
                )
                log_security_event('recaptcha_failed', email, f'IP: {client_ip}, Action: login')
                context['form'] = form
                return render(request, 'auth/login.html', context)
        
        # Check if account or IP is locked
        try:
            email_locked = AccountLockout.is_locked(email) if email else False
            ip_locked = AccountLockout.is_locked(client_ip)
            
            if email_locked or ip_locked:
                remaining_time = max(
                    AccountLockout.get_lockout_time_remaining(email) if email else 0,
                    AccountLockout.get_lockout_time_remaining(client_ip)
                )
                minutes = max(1, remaining_time // 60)
                messages.error(
                    request,
                    f'üîí Account temporarily locked due to too many failed login attempts. '
                    f'Please try again in {minutes} minutes.'
                )
                log_security_event('login_blocked', email, f'IP: {client_ip}')
                context.update({
                    'form': form,
                    'locked': True,
                    'remaining_minutes': minutes
                })
                return render(request, 'auth/login.html', context)
        except Exception as e:
            # If lockout check fails, log and continue (fail open)
            security_logger.error(f"Lockout check failed: {e}")
        
        if form.is_valid():
            user = form.get_user()
            if user:
                # Successful login - reset failed attempts
                try:
                    AccountLockout.reset_failed_attempts(email)
                    AccountLockout.reset_failed_attempts(client_ip)
                except Exception as e:
                    security_logger.error(f"Failed to reset attempts: {e}")
                
                login(request, user)
                messages.success(request, f'Welcome back, {user.first_name or user.email}!')
                log_security_event('login_success', email, f'IP: {client_ip}')
                return redirect('ghg:index')
            else:
                # User not found or inactive
                messages.error(
                    request,
                    '‚ö†Ô∏è Invalid email or password. Please check your credentials.'
                )
        else:
            # Failed login - increment attempts
            if email:
                try:
                    email_attempts = AccountLockout.increment_failed_attempts(email)
                    ip_attempts = AccountLockout.increment_failed_attempts(client_ip)
                    
                    attempts_remaining = AccountLockout.get_attempts_remaining(email)
                    
                    if attempts_remaining > 0:
                        messages.error(
                            request,
                            f'‚ö†Ô∏è Invalid email or password. '
                            f'You have {attempts_remaining} attempts remaining before account lockout.'
                        )
                    
                    log_security_event(
                        'login_failed',
                        email,
                        f'IP: {client_ip}, Attempts: {email_attempts}'
                    )
                except Exception as e:
                    # If increment fails, just show generic error
                    security_logger.error(f"Failed to increment attempts: {e}")
                    messages.error(
                        request,
                        '‚ö†Ô∏è Invalid email or password. Please check your credentials.'
                    )
    else:
        form = EmailLoginForm()
    
    context['form'] = form
    return render(request, 'auth/login.html', context)


@login_required
def security_status(request):
    """Show security status for admin users"""
    if not request.user.is_staff:
        return HttpResponseForbidden("Access denied")
    
    from .security import AccountLockout, get_client_ip
    from django.core.cache import cache
    
    # Get all locked accounts from cache
    locked_accounts = []
    failed_attempts = []
    
    # This is a simplified version - in production, you'd want to store this in DB
    context = {
        'locked_accounts': locked_accounts,
        'failed_attempts': failed_attempts,
        'active_menu': 'security',
    }
    
    return render(request, 'security_status.html', context)


@arcjet_protect()  # Enhanced security with Arcjet simulation
def email_signup_view(request):
    """Signup view using email with reCAPTCHA protection"""
    from .captcha import ReCaptchaValidator, get_recaptcha_context
    from .security import get_client_ip, log_security_event
    
    if request.user.is_authenticated:
        return redirect('ghg:index')
    
    # Add reCAPTCHA context
    context = get_recaptcha_context()
    
    if request.method == 'POST':
        form = EmailSignupForm(request.POST)
        client_ip = get_client_ip(request)
        email = request.POST.get('email', '')
        
        # Validate reCAPTCHA first (if enabled)
        recaptcha_token = request.POST.get('recaptcha_token', '')
        if context.get('RECAPTCHA_ENABLED'):
            if not ReCaptchaValidator.is_human(recaptcha_token, action='signup', remote_ip=client_ip):
                messages.error(
                    request,
                    'ü§ñ Security verification failed. Please try again.'
                )
                log_security_event('recaptcha_failed', email, f'IP: {client_ip}, Action: signup')
                context['form'] = form
                return render(request, 'auth/signup.html', context)
        
        if form.is_valid():
            user = form.save()
            login(request, user, backend='ghg.backends.EmailBackend')
            messages.success(request, 'Account created successfully! Welcome to SustIndex.')
            log_security_event('signup_success', email, f'IP: {client_ip}')
            return redirect('ghg:index')
        else:
            log_security_event('signup_failed', email, f'IP: {client_ip}, Errors: {form.errors}')
    else:
        form = EmailSignupForm()
    
    context['form'] = form
    return render(request, 'auth/signup.html', context)


def logout_view(request):
    """Logout view"""
    logout(request)
    messages.info(request, 'You have been logged out successfully')
    return redirect('ghg:landing')


@login_required
def get_suppliers(request):
    """API endpoint to get user's suppliers"""
    from .models import Supplier
    
    suppliers = Supplier.objects.filter(user=request.user).values('id', 'name', 'supplier_type')
    return JsonResponse({'suppliers': list(suppliers)})


@login_required
@require_http_methods(["POST"])
@ratelimit(key='user', rate='10/m', method='POST', block=True)
def add_supplier_detailed(request):
    """API endpoint to add a new supplier with full details"""
    from .models import Supplier
    import json
    
    data = json.loads(request.body)
    name = data.get('name', '').strip()
    
    if not name:
        return JsonResponse({'error': 'Supplier name is required'}, status=400)
    
    # Check if supplier already exists
    if Supplier.objects.filter(user=request.user, name=name).exists():
        return JsonResponse({'error': 'Supplier with this name already exists'}, status=400)
    
    # Combine phone code and number
    phone_code = data.get('phone_code', '')
    phone_number = data.get('phone', '')
    full_phone = f"{phone_code} {phone_number}".strip() if phone_code and phone_number else phone_number
    
    supplier = Supplier.objects.create(
        user=request.user,
        name=name,
        supplier_type=data.get('supplier_type') or None,
        contact_person=data.get('contact_person') or None,
        email=data.get('email') or None,
        phone=full_phone or None,
        country=data.get('country') or None,
        city=data.get('city') or None,
        tax_number=data.get('tax_number') or None,
        website=data.get('website') or None,
        address=data.get('address') or None,
        notes=data.get('notes') or None
    )
    
    return JsonResponse({
        'success': True,
        'supplier': {
            'id': supplier.id,
            'name': supplier.name,
            'supplier_type': supplier.supplier_type
        }
    })



@login_required
def get_custom_factors(request):
    """API endpoint to get user's custom emission factors"""
    from .models import CustomEmissionFactor
    
    category = request.GET.get('category', None)
    
    factors = CustomEmissionFactor.objects.filter(user=request.user)
    if category:
        factors = factors.filter(category=category)
    
    data = {
        'factors': [
            {
                'id': f.id,
                'material_name': f.material_name,
                'emission_factor': f.emission_factor,
                'unit': f.unit,
                'category': f.category,
                'is_verified': f.is_verified,
                'supplier_name': f.supplier.name if f.supplier else None,
            }
            for f in factors
        ]
    }
    
    return JsonResponse(data)


@login_required
@csrf_protect
@require_http_methods(["POST"])
@ratelimit(key='user', rate='10/m', method='POST', block=True)
@arcjet_protect()  # Enhanced security with Arcjet simulation
def add_custom_factor(request):
    """API endpoint to add a custom emission factor"""
    from .models import CustomEmissionFactor
    from django.db import IntegrityError
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        data = json.loads(request.body)
        
        name = data.get('material_name', '').strip() or data.get('name', '').strip()
        factor_value = data.get('emission_factor') or data.get('factor_value')
        unit = data.get('unit', '').strip()
        category = data.get('category', '').strip()
        
        # Validate required fields
        if not all([name, factor_value, unit, category]):
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        # Validate factor value
        try:
            factor_value = float(factor_value)
            if factor_value < 0:
                return JsonResponse({'error': 'Emission factor cannot be negative'}, status=400)
            if factor_value > 1000:
                return JsonResponse({'error': 'Emission factor seems unreasonably high'}, status=400)
        except ValueError:
            return JsonResponse({'error': 'Invalid emission factor value'}, status=400)
        
        # Limit field lengths for security
        name = name[:200]
        category = category[:100]
        unit = unit[:50]
        description = data.get('description', '').strip()[:2000]
        reference_source = (data.get('source_reference', '').strip() or data.get('reference_source', '').strip())[:500]
        
        # Check if custom factor with same name already exists for this user
        if CustomEmissionFactor.objects.filter(user=request.user, name=name).exists():
            return JsonResponse({
                'error': f'A custom emission factor with the name "{name}" already exists. Please use a different name.'
            }, status=400)
        
        # Create custom factor
        custom_factor = CustomEmissionFactor.objects.create(
            user=request.user,
            name=name,
            category=category,
            description=description,
            factor_value=factor_value,
            unit=unit,
            reference_source=reference_source
        )
        
        logger.info(f"Custom factor created: {name} by user {request.user.id}")
        security_logger.info(f"User {request.user.id} added custom emission factor: {name}")
        
        return JsonResponse({
            'success': True,
            'factor': {
                'id': custom_factor.id,
                'name': custom_factor.name,
                'factor_value': custom_factor.factor_value,
                'unit': custom_factor.unit,
                'category': custom_factor.category,
            }
        })
        
    except json.JSONDecodeError:
        security_logger.warning(f"Invalid JSON from user {request.user.id} in add_custom_factor")
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except IntegrityError as e:
        logger.error(f"IntegrityError in add_custom_factor for user {request.user.id}: {str(e)}")
        return JsonResponse({
            'error': 'A custom emission factor with this name already exists. Please use a different name.'
        }, status=400)
    except Exception as e:
        logger.error(f"Error adding custom factor for user {request.user.id}: {str(e)}")
        security_logger.error(f"Error adding custom factor for user {request.user.id}: {str(e)}")
        return JsonResponse({'error': f'Failed to add custom factor: {str(e)}'}, status=500)


@login_required
@csrf_protect
@require_http_methods(["POST"])
@arcjet_protect()  # Enhanced security with Arcjet simulation
def request_new_material(request):
    """API endpoint to request a new material/source"""
    from .models import MaterialRequest
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        data = json.loads(request.body)
        
        material_name = data.get('material_name', '').strip()
        category = data.get('category', '').strip()
        description = data.get('description', '').strip()
        
        # Validate required fields
        if not all([material_name, category, description]):
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        # Limit field lengths for security
        material_name = material_name[:200]
        category = category[:100]
        description = description[:2000]
        
        # Get optional fields
        suggested_factor = data.get('suggested_factor')
        suggested_unit = data.get('suggested_unit', '').strip()[:50] if data.get('suggested_unit') else None
        suggested_source = data.get('suggested_source', '').strip()[:500] if data.get('suggested_source') else None
        
        # Validate suggested factor if provided
        if suggested_factor:
            try:
                suggested_factor = float(suggested_factor)
                if suggested_factor < 0:
                    return JsonResponse({'error': 'Suggested factor cannot be negative'}, status=400)
            except (ValueError, TypeError):
                suggested_factor = None
        
        # Build additional_info string
        additional_parts = [f"Category: {category}"]
        if suggested_factor:
            factor_str = f"Suggested Factor: {suggested_factor}"
            if suggested_unit:
                factor_str += f" {suggested_unit}"
            additional_parts.append(factor_str)
        if suggested_source:
            additional_parts.append(f"Source: {suggested_source}")
        
        additional_info = "\n".join(additional_parts)
        
        # Create material request
        material_request = MaterialRequest.objects.create(
            user=request.user,
            request_type='material',
            name=material_name,
            description=description,
            additional_info=additional_info
        )
        
        logger.info(f"Material request created: {material_name} by user {request.user.id}")
        security_logger.info(f"User {request.user.id} requested new material: {material_name}")
        
        return JsonResponse({
            'success': True,
            'message': 'Your request has been submitted. Admin will review it soon.',
            'request_id': material_request.id
        })
        
    except json.JSONDecodeError:
        security_logger.warning(f"Invalid JSON from user {request.user.id} in request_new_material")
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error creating material request for user {request.user.id}: {str(e)}")
        security_logger.error(f"Error creating material request for user {request.user.id}: {str(e)}")
        return JsonResponse({'error': 'Failed to submit request'}, status=500)


@login_required
def get_industry_types(request):
    """API endpoint to get available industry types"""
    from .models import IndustryType
    
    industries = IndustryType.objects.filter(is_active=True).order_by('name')
    
    data = {
        'industries': [
            {
                'id': industry.id,
                'name': industry.name,
                'code': industry.code,
                'description': industry.description,
            }
            for industry in industries
        ]
    }
    
    return JsonResponse(data)


@login_required
def request_new_industry(request):
    """API endpoint to request a new industry type"""
    from .models import IndustryRequest
    import json
    import logging
    
    logger = logging.getLogger(__name__)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data provided'}, status=400)
        
        industry_name = data.get('industry_name', '').strip()
        description = data.get('description', '').strip()
        
        # Enhanced validation
        if not industry_name:
            return JsonResponse({'error': 'Industry name is required'}, status=400)
        
        if len(industry_name) < 3:
            return JsonResponse({'error': 'Industry name must be at least 3 characters long'}, status=400)
        
        if len(industry_name) > 200:
            return JsonResponse({'error': 'Industry name cannot exceed 200 characters'}, status=400)
        
        if not description:
            return JsonResponse({'error': 'Industry description is required'}, status=400)
        
        if len(description) < 10:
            return JsonResponse({'error': 'Description must be at least 10 characters long'}, status=400)
        
        if len(description) > 1000:
            return JsonResponse({'error': 'Description cannot exceed 1000 characters'}, status=400)
        
        # Check if industry already exists
        from .models import IndustryType
        if IndustryType.objects.filter(name__iexact=industry_name).exists():
            return JsonResponse({
                'error': f'Industry type "{industry_name}" already exists in our system. Please select it from the dropdown.'
            }, status=400)
        
        # Check if user already requested this industry
        existing_request = IndustryRequest.objects.filter(
            user=request.user, 
            industry_name__iexact=industry_name,
            status='pending'
        ).first()
        
        if existing_request:
            return JsonResponse({
                'error': f'You have already requested "{industry_name}" on {existing_request.created_at.strftime("%Y-%m-%d")}. Please wait for admin review.'
            }, status=400)
        
        # Check industry code if provided
        industry_code = data.get('industry_code') or ''
        industry_code = industry_code.strip() if industry_code else None
        if industry_code and len(industry_code) > 20:
            return JsonResponse({'error': 'Industry code cannot exceed 20 characters'}, status=400)
        
        # Check business context if provided
        business_context = data.get('business_context') or ''
        business_context = business_context.strip() if business_context else None
        if business_context and len(business_context) > 1000:
            return JsonResponse({'error': 'Business context cannot exceed 1000 characters'}, status=400)
        
        # Check for auto-approval patterns
        auto_approve_patterns = [
            'manufacturing', 'retail', 'healthcare', 'education', 'finance', 
            'technology', 'construction', 'transportation', 'energy', 'agriculture'
        ]
        
        should_auto_approve = any(pattern in industry_name.lower() for pattern in auto_approve_patterns)
        
        try:
            # Create industry request
            industry_request = IndustryRequest.objects.create(
                user=request.user,
                industry_name=industry_name,
                industry_code=industry_code,
                description=description,
                business_context=business_context
            )
            
            # Auto-approve if it matches common patterns
            if should_auto_approve:
                # Create the industry type immediately
                industry_type, created = IndustryType.objects.get_or_create(
                    name=industry_name,
                    defaults={
                        'code': industry_code,
                        'description': description,
                        'requested_by': request.user,
                        'is_active': True,
                    }
                )
                
                # Update request status
                industry_request.status = 'approved'
                industry_request.approved_industry = industry_type
                industry_request.admin_notes = 'Auto-approved based on common industry pattern'
                industry_request.save()
                
                logger.info(f"Auto-approved industry request: {industry_name} by {request.user.email}")
                
                return JsonResponse({
                    'success': True,
                    'message': f'Great! "{industry_name}" has been automatically approved and is now available in the dropdown.',
                    'request_id': industry_request.id,
                    'auto_approved': True
                })
            
            logger.info(f"New industry request: {industry_name} by {request.user.email}")
            
            # Send email notification to admin for manual review
            try:
                from .notifications import send_industry_request_notification
                notification_sent = send_industry_request_notification(industry_request)
                if notification_sent:
                    logger.info(f"Admin notification sent for industry request: {industry_name}")
                else:
                    logger.warning(f"Failed to send admin notification for industry request: {industry_name}")
            except Exception as e:
                logger.error(f"Error sending admin notification: {str(e)}")
                # Don't fail the request if notification fails
            
            return JsonResponse({
                'success': True,
                'message': f'Your request for "{industry_name}" has been submitted successfully! Our admin team will review it within 24-48 hours.',
                'request_id': industry_request.id,
                'auto_approved': False
            })
            
        except Exception as e:
            logger.error(f"Error creating industry request: {str(e)}")
            return JsonResponse({
                'error': 'An unexpected error occurred while processing your request. Please try again.'
            }, status=500)
    
    return JsonResponse({'error': 'Only POST method is allowed'}, status=405)


@login_required
def calculate_with_custom_factor(request):
    """Calculate emissions using a custom emission factor"""
    import json
    from .models import EmissionRecord, CustomEmissionFactor
    
    if request.method == 'POST':
        data = json.loads(request.body)
        
        custom_factor_id = data.get('custom_factor_id')
        activity_data = float(data.get('activity_data', 0))
        description = data.get('description', '')
        
        if not custom_factor_id or activity_data <= 0:
            return JsonResponse({'error': 'Invalid data'}, status=400)
        
        try:
            custom_factor = CustomEmissionFactor.objects.get(id=custom_factor_id, user=request.user)
        except CustomEmissionFactor.DoesNotExist:
            return JsonResponse({'error': 'Custom factor not found'}, status=404)
        
        # Calculate emissions
        emissions_kg = activity_data * custom_factor.emission_factor
        emissions_tons = emissions_kg / 1000.0
        
        # Determine scope based on category
        scope = '3'  # Default to Scope 3
        if custom_factor.category in ['stationary', 'mobile', 'fugitive']:
            scope = '1'
        elif custom_factor.category in ['electricity', 'steam-heat']:
            scope = '2'
        
        # Save to database
        record = EmissionRecord.objects.create(
            user=request.user,
            scope=scope,
            category=custom_factor.category,
            source='custom-' + str(custom_factor.id),
            source_name=custom_factor.material_name,
            activity_data=activity_data,
            unit=custom_factor.unit,
            emission_factor=custom_factor.emission_factor,
            emissions_kg=emissions_kg,
            emissions_tons=emissions_tons,
            country='custom',
            reference=f'Custom factor: {custom_factor.source_reference}' if custom_factor.source_reference else 'Custom emission factor',
            description=description,
            supplier=custom_factor.supplier
        )
        
        result = {
            'emissions_kg': round(emissions_kg, 4),
            'emissions_tons': round(emissions_tons, 6),
            'factor': custom_factor.emission_factor,
            'unit': custom_factor.unit,
            'source_name': custom_factor.material_name,
            'activity_data': activity_data,
            'country': 'custom',
            'category': custom_factor.category,
            'source_key': 'custom-' + str(custom_factor.id),
            'reference': custom_factor.source_reference or 'Custom emission factor',
            'record_id': record.id,
            'saved': True,
            'is_custom': True,
            'is_verified': custom_factor.is_verified
        }
        
        return JsonResponse(result)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def get_emission_record(request, record_id):
    """API endpoint to get a single emission record"""
    from .models import EmissionRecord
    
    try:
        record = EmissionRecord.objects.get(id=record_id, user=request.user)
        
        return JsonResponse({
            'id': record.id,
            'activity_data': record.activity_data,
            'unit': record.unit,
            'description': record.description or '',
            'source_name': record.source_name,
            'emissions_kg': record.emissions_kg,
            'emissions_tons': record.emissions_tons,
            'created_at': record.created_at.isoformat()
        })
        
    except EmissionRecord.DoesNotExist:
        return JsonResponse({'error': 'Record not found'}, status=404)


@login_required
def update_emission_record(request, record_id):
    """API endpoint to update an emission record"""
    from .models import EmissionRecord
    import json
    
    if request.method == 'PUT':
        try:
            record = EmissionRecord.objects.get(id=record_id, user=request.user)
            data = json.loads(request.body)
            
            # Get new activity data
            new_activity_data = float(data.get('activity_data', record.activity_data))
            new_description = data.get('description', record.description or '')
            
            if new_activity_data <= 0:
                return JsonResponse({'error': 'Activity data must be greater than 0'}, status=400)
            
            # Recalculate emissions if activity data changed
            if new_activity_data != record.activity_data:
                record.activity_data = new_activity_data
                record.emissions_kg = new_activity_data * record.emission_factor
                record.emissions_tons = record.emissions_kg / 1000.0
            
            # Update description
            record.description = new_description
            record.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Record updated successfully',
                'record': {
                    'id': record.id,
                    'activity_data': record.activity_data,
                    'emissions_kg': record.emissions_kg,
                    'emissions_tons': record.emissions_tons,
                    'description': record.description
                }
            })
            
        except EmissionRecord.DoesNotExist:
            return JsonResponse({'error': 'Record not found'}, status=404)
        except ValueError:
            return JsonResponse({'error': 'Invalid activity data'}, status=400)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
def delete_emission_record(request, record_id):
    """API endpoint to delete an emission record"""
    from .models import EmissionRecord
    
    if request.method == 'DELETE':
        try:
            record = EmissionRecord.objects.get(id=record_id, user=request.user)
            record.delete()
            
            return JsonResponse({
                'success': True,
                'message': 'Record deleted successfully'
            })
            
        except EmissionRecord.DoesNotExist:
            return JsonResponse({'error': 'Record not found'}, status=404)
    
    return JsonResponse({'error': 'Invalid request method'}, status=400)


@login_required
def export_emission_report(request, scope):
    """API endpoint to export emission records as Excel report"""
    from .models import EmissionRecord
    from django.http import HttpResponse
    import io
    from datetime import datetime
    
    try:
        # Try to import openpyxl for Excel export
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return JsonResponse({'error': 'Excel export not available. Please install openpyxl.'}, status=500)
    
    # Filter records based on scope
    records = EmissionRecord.objects.filter(user=request.user).order_by('-created_at')
    
    if scope != 'all':
        scope_number = scope.replace('scope', '')
        records = records.filter(scope=int(scope_number))
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    
    # Set worksheet title based on scope (max 31 chars for Excel)
    scope_titles = {
        'all': 'All Scopes Report',
        'scope1': 'Scope 1 Direct',
        'scope2': 'Scope 2 Electricity', 
        'scope3': 'Scope 3 Indirect'
    }
    ws.title = scope_titles.get(scope, 'Emission Report')
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D7A5F", end_color="2D7A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"SustIndex - {scope_titles.get(scope, 'Emission Report')}"
    title_cell.font = Font(bold=True, size=16, color="2D7A5F")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add generation info
    ws.merge_cells('A2:H2')
    info_cell = ws['A2']
    info_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | User: {request.user.email}"
    info_cell.font = Font(size=10, color="666666")
    info_cell.alignment = Alignment(horizontal="center")
    
    # Add summary statistics
    total_records = records.count()
    total_emissions_kg = sum(record.emissions_kg for record in records)
    total_emissions_tons = total_emissions_kg / 1000.0
    
    ws.merge_cells('A4:H4')
    summary_cell = ws['A4']
    summary_cell.value = f"Summary: {total_records} records | Total Emissions: {total_emissions_kg:.2f} kg CO2e ({total_emissions_tons:.3f} tons CO2e)"
    summary_cell.font = Font(bold=True, size=12, color="2D7A5F")
    summary_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['Date', 'Time', 'Scope', 'Source', 'Activity Data', 'Unit', 'Emissions (kg CO2e)', 'Emissions (tons CO2e)', 'Country', 'Description']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row, record in enumerate(records, 7):
        data = [
            record.created_at.strftime('%Y-%m-%d'),
            record.created_at.strftime('%H:%M:%S'),
            f"Scope {record.scope}",
            record.source_name,
            record.activity_data,
            record.unit,
            round(record.emissions_kg, 3),
            round(record.emissions_tons, 6),
            'Turkey' if record.country == 'turkey' else 'Global',
            record.description or ''
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            
            # Color code by scope
            if col == 3:  # Scope column
                if record.scope == 1:
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                elif record.scope == 2:
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                elif record.scope == 3:
                    cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Set filename
    scope_names = {
        'all': 'All_Scopes',
        'scope1': 'Scope_1_Direct_Emissions',
        'scope2': 'Scope_2_Electricity',
        'scope3': 'Scope_3_Indirect_Emissions'
    }
    
    today = datetime.now().strftime('%Y-%m-%d')
    filename = f"{scope_names.get(scope, 'Emission')}_Report_{today}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response



@login_required
def analysis_index(request):
    """Analysis landing page with cards"""
    return render(request, 'analysis/index.html')

@login_required
def analysis(request):
    """Analysis page view"""
    from .models import EmissionRecord
    from django.db.models import Sum
    
    # Get user's emission statistics
    records = EmissionRecord.objects.filter(user=request.user)
    total_records = records.count()
    total_emissions_kg = records.aggregate(Sum('emissions_kg'))['emissions_kg__sum'] or 0
    total_emissions = total_emissions_kg / 1000.0  # Convert to tons
    
    context = {
        'total_records': total_records,
        'total_emissions': total_emissions,
        'active_menu': 'analysis',  # Set active menu
    }
    
    return render(request, 'analysis.html', context)


@login_required
def analysis_scope_distribution(request):
    """API endpoint for scope distribution data"""
    from .models import EmissionRecord
    from django.db.models import Sum
    
    records = EmissionRecord.objects.filter(user=request.user)
    
    scope1 = records.filter(scope='1').aggregate(Sum('emissions_kg'))['emissions_kg__sum'] or 0
    scope2 = records.filter(scope='2').aggregate(Sum('emissions_kg'))['emissions_kg__sum'] or 0
    scope3 = records.filter(scope='3').aggregate(Sum('emissions_kg'))['emissions_kg__sum'] or 0
    
    return JsonResponse({
        'scope1': round(scope1, 2),
        'scope2': round(scope2, 2),
        'scope3': round(scope3, 2)
    })


@login_required
def analysis_monthly_trends(request):
    """API endpoint for monthly emission trends"""
    from .models import EmissionRecord
    from django.db.models import Sum
    from django.db.models.functions import TruncMonth
    from datetime import datetime, timedelta
    
    # Get last 12 months of data
    end_date = datetime.now()
    start_date = end_date - timedelta(days=365)
    
    records = EmissionRecord.objects.filter(
        user=request.user,
        created_at__gte=start_date
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        total=Sum('emissions_kg')
    ).order_by('month')
    
    months = []
    emissions = []
    
    for record in records:
        months.append(record['month'].strftime('%b %Y'))
        emissions.append(round(record['total'], 2))
    
    return JsonResponse({
        'months': months,
        'emissions': emissions
    })


@login_required
def analysis_top_sources(request):
    """API endpoint for top emission sources"""
    from .models import EmissionRecord
    from django.db.models import Sum
    
    # Get top 10 emission sources
    sources = EmissionRecord.objects.filter(
        user=request.user
    ).values('source_name').annotate(
        total=Sum('emissions_kg')
    ).order_by('-total')[:10]
    
    # Calculate total for percentage
    total_emissions = sum(s['total'] for s in sources)
    
    sources_data = []
    for source in sources:
        percentage = (source['total'] / total_emissions * 100) if total_emissions > 0 else 0
        sources_data.append({
            'name': source['source_name'],
            'emissions': round(source['total'], 2),
            'percentage': round(percentage, 1)
        })
    
    return JsonResponse({
        'sources': sources_data
    })


@login_required
def emissions(request):
    """Professional emissions analysis page"""
    from datetime import datetime, timedelta
    
    # Set default date range (last 12 months)
    today = datetime.now()
    last_year = today - timedelta(days=365)
    
    context = {
        'date_from': last_year.strftime('%Y-%m-%d'),
        'date_to': today.strftime('%Y-%m-%d'),
        'active_menu': 'emissions',  # Set active menu
    }
    
    return render(request, 'analysis/emissions.html', context)


@login_required
def emissions_data_api(request):
    """API endpoint for emissions data with filters"""
    from .models import EmissionRecord
    from django.db.models import Sum, Q
    from datetime import datetime
    import json
    
    # Get filter parameters
    method = request.GET.get('standard', 'ghg')
    date_from = request.GET.get('from')
    date_to = request.GET.get('to')
    scopes = request.GET.get('scopes', '').split(',') if request.GET.get('scopes') else []
    categories = request.GET.get('categories', '').split(',') if request.GET.get('categories') else []
    sources = request.GET.get('sources', '').split(',') if request.GET.get('sources') else []
    
    # Base queryset
    records = EmissionRecord.objects.filter(user=request.user)
    
    # Apply date filter
    if date_from:
        try:
            records = records.filter(created_at__gte=datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            records = records.filter(created_at__lte=datetime.strptime(date_to, '%Y-%m-%d'))
        except ValueError:
            pass
    
    # Apply scope filter
    if scopes and scopes != ['']:
        try:
            scope_numbers = [int(s) for s in scopes if s.isdigit()]
            if scope_numbers:
                records = records.filter(scope__in=scope_numbers)
        except ValueError:
            pass
    
    # Calculate scope totals
    scope1_total = records.filter(scope='1').aggregate(Sum('emissions_kg'))['emissions_kg__sum'] or 0
    scope2_total = records.filter(scope='2').aggregate(Sum('emissions_kg'))['emissions_kg__sum'] or 0
    scope3_total = records.filter(scope='3').aggregate(Sum('emissions_kg'))['emissions_kg__sum'] or 0
    
    # Get sources data
    sources_data = records.values('source_name', 'scope').annotate(
        total=Sum('emissions_kg')
    ).order_by('-total')
    
    sources_list = []
    total_emissions = scope1_total + scope2_total + scope3_total
    
    for source in sources_data:
        emissions = source['total']
        percentage = (emissions / total_emissions * 100) if total_emissions > 0 else 0
        
        sources_list.append({
            'name': source['source_name'],
            'scope': source['scope'],
            'emissions': emissions,
            'percentage': percentage
        })
    
    # Prepare chart data
    chart_labels = []
    chart_values = []
    
    if scope1_total > 0:
        chart_labels.append('Scope 1')
        chart_values.append(scope1_total)
    if scope2_total > 0:
        chart_labels.append('Scope 2') 
        chart_values.append(scope2_total)
    if scope3_total > 0:
        chart_labels.append('Scope 3')
        chart_values.append(scope3_total)
    
    return JsonResponse({
        'scope1': scope1_total,
        'scope2': scope2_total,
        'scope3': scope3_total,
        'sources': sources_list,
        'scope_chart': {
            'labels': chart_labels,
            'values': chart_values
        },
        'method': method,
        'total_records': records.count(),
        'tree': {
            'scopes': [
                {
                    'id': '1',
                    'name': 'Scope 1',
                    '_collapsed': False,
                    'categories': [
                        {
                            'key': 'stationary',
                            'name': 'Stationary Combustion',
                            '_collapsed': False,
                            'sources': [
                                {'key': 'coal-industrial', 'name': 'Coal (industrial)'},
                                {'key': 'natural-gas', 'name': 'Natural Gas'},
                                {'key': 'diesel-oil', 'name': 'Gas/Diesel Oil'},
                                {'key': 'lpg', 'name': 'LPG'},
                                {'key': 'propane', 'name': 'Propane'},
                                {'key': 'motor-gasoline', 'name': 'Motor Gasoline'}
                            ]
                        },
                        {
                            'key': 'mobile',
                            'name': 'Mobile Combustion',
                            '_collapsed': False,
                            'sources': [
                                {'key': 'off-road-gasoline', 'name': 'Off-Road Gasoline'},
                                {'key': 'off-road-diesel', 'name': 'Off-Road Diesel'},
                                {'key': 'on-road-diesel', 'name': 'On-Road Diesel'},
                                {'key': 'car-gasoline', 'name': 'Car - Gasoline'},
                                {'key': 'car-diesel', 'name': 'Car - Diesel'}
                            ]
                        },
                        {
                            'key': 'fugitive',
                            'name': 'Fugitive Emissions',
                            '_collapsed': False,
                            'sources': [
                                {'key': 'r432a', 'name': 'R432A'},
                                {'key': 'r410a', 'name': 'R410A'},
                                {'key': 'r134a', 'name': 'R134a'}
                            ]
                        }
                    ]
                },
                {
                    'id': '2',
                    'name': 'Scope 2',
                    '_collapsed': False,
                    'categories': [
                        {
                            'key': 'electricity',
                            'name': 'Electricity',
                            '_collapsed': False,
                            'sources': [
                                {'key': 'grid-electricity', 'name': 'Grid Electricity'},
                                {'key': 'renewable-energy', 'name': 'Renewable Energy'}
                            ]
                        },
                        {
                            'key': 'heat-steam',
                            'name': 'Heat and Steam',
                            '_collapsed': False,
                            'sources': [
                                {'key': 'district-heating', 'name': 'District Heating'},
                                {'key': 'steam', 'name': 'Steam'}
                            ]
                        }
                    ]
                },
                {
                    'id': '3',
                    'name': 'Scope 3',
                    '_collapsed': False,
                    'categories': [
                        {
                            'key': 'purchased-goods',
                            'name': 'Purchased Goods',
                            '_collapsed': False,
                            'sources': [
                                {'key': 'paper', 'name': 'Paper'},
                                {'key': 'plastic', 'name': 'Plastic'},
                                {'key': 'metal', 'name': 'Metal'},
                                {'key': 'chemical', 'name': 'Chemical'},
                                {'key': 'wood', 'name': 'Wood'}
                            ]
                        },
                        {
                            'key': 'business-travel',
                            'name': 'Business Travel',
                            '_collapsed': False,
                            'sources': [
                                {'key': 'flight', 'name': 'Flight'},
                                {'key': 'hotel', 'name': 'Hotel'},
                                {'key': 'car-rental', 'name': 'Car Rental'},
                                {'key': 'train', 'name': 'Train'}
                            ]
                        },
                        {
                            'key': 'waste',
                            'name': 'Waste',
                            '_collapsed': False,
                            'sources': [
                                {'key': 'waste', 'name': 'Waste'},
                                {'key': 'recycling', 'name': 'Recycling'},
                                {'key': 'landfill', 'name': 'Landfill'}
                            ]
                        }
                    ]
                }
            ]
        }
    })


@login_required
def emissions_export_api(request):
    """Export emissions data as Excel"""
    from .models import EmissionRecord
    from django.http import HttpResponse
    from datetime import datetime
    import json
    import io
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return JsonResponse({'error': 'Excel export not available'}, status=500)
    
    # Get same filters as data API
    method = request.GET.get('method', 'ghg')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    scopes = request.GET.get('scopes', '1,2,3').split(',')
    
    # Base queryset
    records = EmissionRecord.objects.filter(user=request.user)
    
    # Apply filters
    if date_from:
        records = records.filter(created_at__gte=datetime.strptime(date_from, '%Y-%m-%d'))
    if date_to:
        records = records.filter(created_at__lte=datetime.strptime(date_to, '%Y-%m-%d'))
    if scopes:
        scope_numbers = [int(s) for s in scopes if s.isdigit()]
        records = records.filter(scope__in=scope_numbers)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Emissions Analysis"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Emissions Analysis Report - {method.upper()} Method"
    title_cell.font = Font(bold=True, size=16, color="10B981")
    title_cell.alignment = Alignment(horizontal="center")
    
    # Date range
    ws.merge_cells('A2:F2')
    date_cell = ws['A2']
    date_cell.value = f"Period: {date_from} to {date_to}"
    date_cell.font = Font(size=12)
    date_cell.alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['Source', 'Scope', 'Activity Data', 'Unit', 'Emissions (kg CO2e)', 'Emissions (tCO2e)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, record in enumerate(records.order_by('-emissions_kg'), 5):
        data = [
            record.source_name,
            f"Scope {record.scope}",
            record.activity_data,
            record.unit,
            round(record.emissions_kg, 3),
            round(record.emissions_tons, 6)
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            
            # Color code by scope
            if col == 2:  # Scope column
                if record.scope == 1:
                    cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                elif record.scope == 2:
                    cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
                elif record.scope == 3:
                    cell.fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    today = datetime.now().strftime('%Y-%m-%d')
    filename = f"Emissions_Analysis_{method.upper()}_{today}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

# New professional pages
@login_required
def action_planning(request):
    """Action Planning page view"""
    context = {
        'active_menu': 'action_planning',
    }
    return render(request, 'action_planning.html', context)

@login_required
def suppliers(request):
    """Supplier Management page view"""
    from .models import Supplier
    
    # Get supplier statistics
    user_suppliers = Supplier.objects.filter(user=request.user)
    supplier_count = user_suppliers.count()
    # Since there's no is_active field, we'll count all suppliers as active
    active_suppliers = supplier_count
    supplier_categories = user_suppliers.values('supplier_type').distinct().count()
    countries_count = user_suppliers.values('country').distinct().count()
    
    context = {
        'active_menu': 'suppliers',
        'supplier_count': supplier_count,
        'active_suppliers': active_suppliers,
        'supplier_categories': supplier_categories,
        'countries_count': countries_count,
    }
    return render(request, 'suppliers.html', context)

@login_required
def settings(request):
    """Settings page view"""
    context = {
        'active_menu': 'settings',
    }
    return render(request, 'settings.html', context)

@login_required
def support(request):
    """Support page view"""
    context = {
        'active_menu': 'support',
    }
    return render(request, 'support.html', context)


# REMOVED: fix_users_temp endpoint for security reasons

@login_required
def inventory_report(request):
    """Inventory reporting page with filtering and PDF generation"""
    from datetime import datetime, date
    from django.db.models import Sum, Count, Q
    from .models import EmissionRecord, ReportExtraInfo
    
    # Get filter parameters
    date_from = request.GET.get('from')
    date_to = request.GET.get('to')
    scope_filter = request.GET.get('scope', 'all')
    country_filter = request.GET.get('country', 'all')
    
    # Base queryset
    records = EmissionRecord.objects.filter(user=request.user)
    
    # Apply filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            records = records.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            date_from = None
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            records = records.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            date_to = None
    
    if scope_filter != 'all':
        records = records.filter(scope=scope_filter)
    
    if country_filter != 'all':
        records = records.filter(country=country_filter)
    
    # Calculate summary data
    total_emissions_kg = records.aggregate(total=Sum('emissions_kg'))['total'] or 0
    total_emissions_tons = total_emissions_kg / 1000
    total_records = records.count()
    
    # Scope breakdown
    scope_breakdown = []
    for scope in [1, 2, 3]:
        scope_records = records.filter(scope=scope)
        scope_kg = scope_records.aggregate(total=Sum('emissions_kg'))['total'] or 0
        scope_tons = scope_kg / 1000
        scope_count = scope_records.count()
        
        percentage = (scope_tons / total_emissions_tons * 100) if total_emissions_tons > 0 else 0
        
        scope_breakdown.append({
            'scope': scope,
            'emissions_tons': round(scope_tons, 3),
            'emissions_kg': scope_kg,
            'records_count': scope_count,
            'percentage': round(percentage, 1)
        })
    
    # Top emission sources
    top_sources = (
        records.values('source_name', 'scope', 'category')
        .annotate(
            total_kg=Sum('emissions_kg'),
            records_count=Count('id')
        )
        .order_by('-total_kg')[:10]
    )
    
    for source in top_sources:
        source['total_tons'] = round(source['total_kg'] / 1000, 3)
        source['percentage'] = round((source['total_kg'] / total_emissions_kg * 100), 1) if total_emissions_kg > 0 else 0
    
    # Get report extra info if exists
    try:
        report_extra = ReportExtraInfo.objects.get(user=request.user)
    except ReportExtraInfo.DoesNotExist:
        report_extra = None
    
    # Countries for filter
    countries = records.values_list('country', flat=True).distinct()
    
    context = {
        'total_emissions_tons': round(total_emissions_tons, 3),
        'total_emissions_kg': total_emissions_kg,
        'total_records': total_records,
        'scope_breakdown': scope_breakdown,
        'top_sources': top_sources,
        'report_extra': report_extra,
        'countries': countries,
        'date_from': date_from,
        'date_to': date_to,
        'scope_filter': scope_filter,
        'country_filter': country_filter,
        'active_menu': 'reporting',
    }
    
    return render(request, 'reporting/inventory.html', context)


@login_required
def generate_pdf_report(request):
    """Generate PDF report for emissions inventory"""
    from django.http import HttpResponse
    from datetime import datetime, date
    from django.db.models import Sum, Count
    from .models import EmissionRecord, ReportExtraInfo
    import io
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        return JsonResponse({'error': 'PDF generation not available. Please install reportlab.'}, status=500)
    
    # Get filter parameters (same as inventory_report)
    date_from = request.GET.get('from')
    date_to = request.GET.get('to')
    scope_filter = request.GET.get('scope', 'all')
    country_filter = request.GET.get('country', 'all')
    
    # Base queryset
    records = EmissionRecord.objects.filter(user=request.user)
    
    # Apply filters
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            records = records.filter(created_at__date__gte=date_from_obj)
        except ValueError:
            date_from = None
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            records = records.filter(created_at__date__lte=date_to_obj)
        except ValueError:
            date_to = None
    
    if scope_filter != 'all':
        records = records.filter(scope=scope_filter)
    
    if country_filter != 'all':
        records = records.filter(country=country_filter)
    
    # Calculate data
    total_emissions_kg = records.aggregate(total=Sum('emissions_kg'))['total'] or 0
    total_emissions_tons = total_emissions_kg / 1000
    
    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        textColor=colors.HexColor('#1e293b')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.HexColor('#374151')
    )
    
    # Title
    title = Paragraph("GHG Emissions Inventory Report", title_style)
    elements.append(title)
    
    # Organization info
    org_info = f"""
    <b>Organization:</b> {request.user.username.title()}<br/>
    <b>Report Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>
    <b>Reporting Standard:</b> ISO 14064-1<br/>
    <b>Report Period:</b> {date_from or 'All time'} to {date_to or 'Present'}
    """
    elements.append(Paragraph(org_info, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Executive Summary
    elements.append(Paragraph("Executive Summary", heading_style))
    summary_text = f"""
    This report presents the greenhouse gas (GHG) emissions inventory for {request.user.username.title()} 
    in accordance with ISO 14064-1 standards. The total emissions for the reporting period are 
    <b>{total_emissions_tons:.3f} tCO‚ÇÇe</b> across {records.count()} emission sources.
    """
    elements.append(Paragraph(summary_text, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Scope breakdown table
    elements.append(Paragraph("Emissions by Scope", heading_style))
    
    scope_data = [['Scope', 'Description', 'Emissions (tCO‚ÇÇe)', 'Percentage']]
    for scope in [1, 2, 3]:
        scope_records = records.filter(scope=scope)
        scope_kg = scope_records.aggregate(total=Sum('emissions_kg'))['total'] or 0
        scope_tons = scope_kg / 1000
        percentage = (scope_tons / total_emissions_tons * 100) if total_emissions_tons > 0 else 0
        
        descriptions = {
            1: 'Direct emissions from owned sources',
            2: 'Indirect emissions from purchased energy',
            3: 'Other indirect emissions in value chain'
        }
        
        scope_data.append([
            f'Scope {scope}',
            descriptions[scope],
            f'{scope_tons:.3f}',
            f'{percentage:.1f}%'
        ])
    
    scope_table = Table(scope_data)
    scope_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb'))
    ]))
    
    elements.append(scope_table)
    elements.append(Spacer(1, 20))
    
    # Top sources
    elements.append(Paragraph("Top Emission Sources", heading_style))
    
    top_sources = (
        records.values('source_name', 'scope', 'category')
        .annotate(total_kg=Sum('emissions_kg'))
        .order_by('-total_kg')[:10]
    )
    
    sources_data = [['Source', 'Scope', 'Category', 'Emissions (tCO‚ÇÇe)', 'Percentage']]
    for source in top_sources:
        source_tons = source['total_kg'] / 1000
        percentage = (source['total_kg'] / total_emissions_kg * 100) if total_emissions_kg > 0 else 0
        
        sources_data.append([
            source['source_name'],
            f"Scope {source['scope']}",
            source['category'].replace('-', ' ').title(),
            f'{source_tons:.3f}',
            f'{percentage:.1f}%'
        ])
    
    sources_table = Table(sources_data)
    sources_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb'))
    ]))
    
    elements.append(sources_table)
    elements.append(Spacer(1, 20))
    
    # Methodology
    elements.append(Paragraph("Methodology", heading_style))
    methodology_text = """
    This inventory was prepared following ISO 14064-1:2018 guidelines for greenhouse gas inventories. 
    Emission factors were sourced from internationally recognized databases including IPCC, Defra, EPA, 
    and country-specific sources. All calculations follow the operational control approach for 
    organizational boundary definition.
    """
    elements.append(Paragraph(methodology_text, styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="emissions_inventory_{datetime.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)
    
    return response

def landing_page(request):
    """ÿµŸÅÿ≠Ÿá ŸÑŸÜÿØ€åŸÜ⁄Ø ÿßÿµŸÑ€å ÿ≥ÿß€åÿ™"""
    # ÿß⁄Øÿ± ⁄©ÿßÿ±ÿ®ÿ± ŸÑÿß⁄Ø€åŸÜ ⁄©ÿ±ÿØŸá ÿ®ÿßÿ¥ÿØÿå ÿ®Ÿá ÿØÿßÿ¥ÿ®Ÿàÿ±ÿØ ŸáÿØÿß€åÿ™ ÿ¥ŸàÿØ
    if request.user.is_authenticated:
        return redirect('ghg:index')
    
    context = {
        'page_title': 'SustIndex - Carbon Tracking Platform',
        'meta_description': 'Professional carbon emission tracking and reporting platform for organizations. ISO 14064-1 compliant reporting, real-time analytics, and comprehensive carbon management tools.',
    }
    return render(request, 'landing.html', context)


def landing_nature(request):
    """ÿµŸÅÿ≠Ÿá ŸÑŸÜÿØ€åŸÜ⁄Ø ÿ®ÿß ÿ∑ÿ±ÿßÿ≠€å ÿ∑ÿ®€åÿπ€å"""
    # ÿß⁄Øÿ± ⁄©ÿßÿ±ÿ®ÿ± ŸÑÿß⁄Ø€åŸÜ ⁄©ÿ±ÿØŸá ÿ®ÿßÿ¥ÿØÿå ÿ®Ÿá ÿØÿßÿ¥ÿ®Ÿàÿ±ÿØ ŸáÿØÿß€åÿ™ ÿ¥ŸàÿØ
    if request.user.is_authenticated:
        return redirect('ghg:index')
    
    context = {
        'page_title': 'Carbon Track - Sustainable Carbon Management Solutions',
        'meta_description': 'World-class carbon tracking platform building sustainable solutions for modern businesses. Tailored solutions for every stage of your sustainability journey.',
    }
    return render(request, 'landing_nature.html', context)


@login_required
def test_language(request):
    """Test page for language switching"""
    return render(request, 'test_language.html')


@login_required
def test_translation(request):
    """Test view for translation functionality"""
    return render(request, 'test_translation.html')

def test_lang_switch(request):
    """Test page for language switching"""
    return render(request, 'test_lang_switch.html', {
        'active_menu': 'test'
    })